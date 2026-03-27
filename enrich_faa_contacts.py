"""
enrich_faa_contacts.py
----------------------
Enriches the HangarLinks_Master_List.xlsx with owner and manager contact
information by querying AirNav (which reflects FAA NASR data) and attempts
to find emails via web search.

PHASE 1 (AirNav): Looks up the exact FAA ID to get Owner Name, Addr, Phone
PHASE 2 (Web Search): If the facility has a likely website, uses DuckDuckGo
                      to find it and scrapes for target emails.

Output:  HangarLinks_Contacts.xlsx
         - One row per facility
         - Fully enriched with all available owner/manager information.

Dependencies:
    pip install requests openpyxl beautifulsoup4 lxml
"""

import io
import re
import time
import requests
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup
from urllib.parse import urlparse

# ── Config ─────────────────────────────────────────────────────────────────────

INPUT_FILE   = "HangarLinks_Master_List.xlsx"
OUTPUT_FILE  = "HangarLinks_Contacts.xlsx"

REQUEST_DELAY = 1.0      # Be polite to AirNav and DuckDuckGo

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    )
}

EMAIL_REGEX = re.compile(
    r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}"
)

SKIP_DOMAINS = {
    "example.com", "gmail.com", "yahoo.com", "hotmail.com",
    "wix.com", "squarespace.com", "godaddy.com", "sentry.io",
    "google.com", "cloudflare.com", "faa.gov",
}

# ── Helpers ────────────────────────────────────────────────────────────────────

def find_emails_in_html(html: str, base_domain: str = "") -> list[str]:
    candidates = EMAIL_REGEX.findall(html)
    seen = set()
    results = []
    for email in candidates:
        email = email.lower().strip(".,;")
        domain = email.split("@")[-1]
        if domain in SKIP_DOMAINS:
            continue
        if base_domain and domain == base_domain:
            if email not in seen:
                seen.add(email)
                results.insert(0, email)
        else:
            if email not in seen:
                seen.add(email)
                results.append(email)
    return results

def fetch_page(url: str, timeout: int = 15) -> str | None:
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
        if r.ok:
            return r.text
    except Exception:
        pass
    return None

def google_search_first_url(query: str) -> str | None:
    try:
        url = "https://html.duckduckgo.com/html/"
        data = {"q": query, "kl": "us-en"}
        r = requests.post(url, data=data, headers=HEADERS, timeout=10)
        if not r.ok:
            return None
        soup = BeautifulSoup(r.text, "lxml")
        for a in soup.select("a.result__url"):
            href = a.get("href", "")
            if href and not href.startswith("https://duckduckgo"):
                return href
        for a in soup.select(".result__a"):
            href = a.get("href", "")
            if "uddg=" in href:
                from urllib.parse import unquote, parse_qs
                parsed = urlparse(href)
                qs = parse_qs(parsed.query)
                if "uddg" in qs:
                    return unquote(qs["uddg"][0])
    except Exception:
        pass
    return None

# ── Feature Extraction ────────────────────────────────────────────────────────

def fetch_airnav_owner_data(faa_id: str) -> dict:
    """Scrapes AirNav for the FAA ownership record given an FAA ID."""
    info = {
        "owner_name": "", "owner_addr": "", "owner_city": "", "owner_state": "", "owner_zip": "", "owner_phone": "",
        "manager_name": "", "manager_phone": ""
    }
    html = fetch_page(f"https://www.airnav.com/airport/{faa_id}", timeout=10)
    if not html:
        return info

    soup = BeautifulSoup(html, "lxml")
    for td in soup.find_all("td"):
        text = td.text.strip()
        if text == "Owner:":
            sib = td.find_next_sibling("td")
            if sib:
                lines = [line.strip() for line in sib.stripped_strings]
                if lines:
                    info["owner_name"] = lines[0]
                if len(lines) >= 3 and "Phone" not in lines[1]:
                    info["owner_addr"] = lines[1]
                    city_state_zip = lines[2].split(", ")
                    if len(city_state_zip) == 2:
                        info["owner_city"] = city_state_zip[0]
                        st_zip = city_state_zip[1].split()
                        if len(st_zip) >= 2:
                            info["owner_state"] = st_zip[0]
                            info["owner_zip"] = st_zip[1]
                for line in lines:
                    if line.startswith("Phone"):
                        info["owner_phone"] = line.replace("Phone", "").strip()
        elif text == "Manager:":
            sib = td.find_next_sibling("td")
            if sib:
                lines = [line.strip() for line in sib.stripped_strings]
                if lines:
                    info["manager_name"] = lines[0]
                for line in lines:
                    if line.startswith("Phone"):
                        info["manager_phone"] = line.replace("Phone", "").strip()
    return info

def find_email_for_facility(name: str, city: str, state: str) -> tuple[str, str, str]:
    query = f'"{name}" {city} {state} airport airpark contact email'
    website = google_search_first_url(query)
    if not website: return ("", "", "not found")

    if not website.startswith("http"): website = "https://" + website
    parsed = urlparse(website)
    base_domain = parsed.netloc.replace("www.", "")

    time.sleep(REQUEST_DELAY)
    html = fetch_page(website)
    emails = find_emails_in_html(html or "", base_domain)
    if emails: return (website, emails[0], "homepage")

    for path in ["/contact", "/contact-us", "/about"]:
        contact_url = f"{parsed.scheme}://{parsed.netloc}{path}"
        time.sleep(REQUEST_DELAY)
        html = fetch_page(contact_url)
        if html:
            emails = find_emails_in_html(html, base_domain)
            if emails: return (website, emails[0], f"contact page ({path})")

    return (website, "", "website found, no email")

# ── Main Process ──────────────────────────────────────────────────────────────

def enrich_facilities(facilities: list[dict], do_email_search: bool) -> list[dict]:
    print(f"\n[Enrichment] Processing {len(facilities)} facilities…")
    if do_email_search:
        print("  Running FULL process (AirNav public records + Email web search)")
    else:
        print("  Running FAST process (AirNav public records only)")
    print("  Press Ctrl+C at any time to save progress and exit gracefully.\n")

    for i, row in enumerate(facilities, 1):
        name = row.get("facility_name", "")
        city = row.get("city", "")
        state = row.get("state", "")
        faa_id = row.get("faa_id", "")
        print(f"  [{i:>3}/{len(facilities)}] {faa_id:<4} | {name[:25]:<25} | {city[:15]:<15}", end=" … ", flush=True)

        try:
            # 1. Gather NASR public ownership from Airnav
            if faa_id:
                owner_data = fetch_airnav_owner_data(faa_id)
                row.update(owner_data)
                time.sleep(REQUEST_DELAY)
            
            # 2. Extract Web/Email if enabled
            if do_email_search:
                website, email, source = find_email_for_facility(name, city, state)
                row["website"] = website
                row["email"] = email
                row["email_source"] = source
                time.sleep(REQUEST_DELAY)
            else:
                row["website"] = ""
                row["email"] = ""
                row["email_source"] = "not searched"
            
            # Display result summary back to terminal
            icons = []
            if row.get("owner_name"): icons.append("👤")
            if row.get("owner_phone"): icons.append("📞")
            if row.get("email"): icons.append("✉️")
            print(" ".join(icons) if icons else "—")
            
        except KeyboardInterrupt:
            print("\n  ⚠ Interrupted! Saving what we found so far...")
            for remaining in facilities[i-1:]:
                remaining.update({
                    "owner_name": "", "owner_addr": "", "owner_city": "", "owner_state": "", "owner_zip": "", "owner_phone": "",
                    "manager_name": "", "manager_phone": "", "website": "", "email": "", "email_source": "skipped"
                })
            break
        except Exception as e:
            print(f"Error ({e})")
    return facilities

# ── Excel Writer ───────────────────────────────────────────────────────────────

def write_contacts_excel(rows: list[dict], filepath: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HangarLinks Contacts"

    COLUMNS = [
        ("facility_name",  "Facility Name"),
        ("city",           "City"),
        ("state",          "State"),
        ("faa_id",         "FAA Airport ID"),
        ("owner_name",     "Owner Name"),
        ("owner_phone",    "Owner Phone"),
        ("owner_addr",     "Owner Address"),
        ("owner_city",     "Owner City"),
        ("owner_state",    "Owner State"),
        ("owner_zip",      "Owner ZIP"),
        ("manager_name",   "Manager Name"),
        ("manager_phone",  "Manager Phone"),
        ("website",        "Website"),
        ("email",          "Email (found)"),
        ("email_source",   "Email Source"),
        ("use_type",       "Use Type"),
        ("ownership_type", "Ownership Type"),
        ("status",         "Status"),
        ("hangar_storage", "Hangar Storage (Y/N)"),
        ("latitude",       "Latitude"),
        ("longitude",      "Longitude"),
        ("elevation",      "Elevation (ft)"),
        ("icao_id",        "ICAO ID"),
    ]

    header_fill = PatternFill("solid", fgColor="1A3A5C")
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    email_fill = PatternFill("solid", fgColor="D4EDDA")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Calibri", size=10)
    email_font = Font(name="Calibri", size=10, color="155724", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Border(left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"), top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"))

    for col_i, (_, label) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=col_i, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin
    ws.row_dimensions[1].height = 28

    for row_i, row in enumerate(rows, 2):
        use_alt = row_i % 2 == 0
        has_email = bool(row.get("email"))
        for col_i, (key, _) in enumerate(COLUMNS, 1):
            val = row.get(key, "") or ""
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.border = thin
            if key == "email" and has_email:
                c.font = email_font
                c.fill = email_fill
                c.alignment = left
            elif key == "website" and val:
                c.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
                c.alignment = left
            else:
                c.font = body_font
                c.alignment = left if col_i <= 2 else center
                if use_alt: c.fill = alt_fill

    for col_i, (_, label) in enumerate(COLUMNS, 1):
        col_letter = get_column_letter(col_i)
        ws.column_dimensions[col_letter].width = min(len(label) + 12, 45)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet("Statistics")
    import datetime
    total = len(rows)
    with_email = sum(1 for r in rows if r.get("email"))
    with_website = sum(1 for r in rows if r.get("website"))
    with_owner = sum(1 for r in rows if r.get("owner_name"))
    with_phone = sum(1 for r in rows if r.get("owner_phone"))

    for r_i, (lvl, val) in enumerate([
        ("HangarLinks Contacts", ""),
        ("Generated", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Total Facilities", total),
        ("With Owner Info", f"{with_owner} ({with_owner*100//total}%)" if total else 0),
        ("With Phone Info", f"{with_phone} ({with_phone*100//total}%)" if total else 0),
        ("With Email Info", f"{with_email} ({with_email*100//total}%)" if total else 0),
    ], 1):
        ws2.cell(row=r_i, column=1, value=lvl)
        ws2.cell(row=r_i, column=2, value=val)
    
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 30
    wb.save(filepath)
    print(f"\n  ✅ Saved → {filepath}")
    print(f"  📊 {with_email} Emails | {with_owner} Owners | {with_phone} Phones")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("  HangarLinks — Unified Content & Owner Enrichment Pipeline")
    print("=" * 65)

    try:
        wb = load_workbook(INPUT_FILE)
        ws = wb["HangarLinks Facilities"]
    except Exception:
        ws = load_workbook(INPUT_FILE).active

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_map = {"FAA Airport ID": "faa_id", "Facility Name": "facility_name", "City": "city", "State": "state", "Ownership Type": "ownership_type", "Use Type": "use_type", "Status": "status", "Latitude": "latitude", "Longitude": "longitude", "Elevation (ft)": "elevation", "ICAO ID": "icao_id", "Hangar Storage (Y/N)": "hangar_storage"}
    
    facilities = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = {}
        for col_i, val in enumerate(row):
            hdr = headers[col_i] if col_i < len(headers) else None
            key = col_map.get(hdr, hdr)
            if key: rec[key] = val or ""
        if rec and rec.get("faa_id"): facilities.append(rec)

    print(f"  ✓ {len(facilities)} facilities loaded.")

    ans = input("  Run FULL enrichment (Web Search + Airnav) (~10 mins)? [Y/n]: ").strip().lower()
    do_email = ans in ("y", "yes", "")

    facilities = enrich_facilities(facilities, do_email)
    write_contacts_excel(facilities, OUTPUT_FILE)

if __name__ == "__main__":
    main()
