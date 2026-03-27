"""
fetch_faa_facilities.py
-----------------------
Downloads the full NTAD Aviation Facilities dataset from the FAA's ArcGIS
Open Data API, filters for Private-use (PR) facilities that contain the
keywords 'Airpark' or 'Hangar' in the facility name, and exports the
results to HangarLinks_Master_List.xlsx.

Usage:
    python fetch_faa_facilities.py

Dependencies:
    pip install requests openpyxl
"""

import requests
import time
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────────

# FAA NTAD Aviation Facilities – ArcGIS REST endpoint
BASE_URL = (
    "https://services.arcgis.com/xOi1kZaI0eWDREZv/arcgis/rest/services/"
    "NTAD_Aviation_Facilities/FeatureServer/0/query"
)

PAGE_SIZE      = 1_000          # Records to fetch per request
KEYWORDS       = ["airpark", "hangar"]  # Case-insensitive keyword filter
PRIVATE_CODE   = "PR"           # FAA "USE" field value for private airports
OUTPUT_FILE    = "HangarLinks_Master_List.xlsx"
REQUEST_DELAY  = 0.5            # Seconds between requests (be polite)

# Columns we care about from the raw API response
# Field names verified against the live ArcGIS schema (2025-12-25 effective date)
FIELD_MAP = {
    "SITE_NO":             "Site Number",
    "ARPT_ID":             "FAA Airport ID",
    "ARPT_NAME":           "Facility Name",
    "CITY":                "City",
    "STATE_CODE":          "State",
    "COUNTY_NAME":         "County",
    "OWNERSHIP_TYPE_CODE": "Ownership Type",
    "FACILITY_USE_CODE":   "Use Type",
    "ARPT_STATUS":         "Status",
    "LAT_DECIMAL":         "Latitude",
    "LONG_DECIMAL":        "Longitude",
    "ELEV":                "Elevation (ft)",
    "ICAO_ID":             "ICAO ID",
    "CHART_NAME":          "Sectional Chart",
    "FUEL_TYPES":          "Fuel Types",
    "TRNS_STRG_HGR_FLAG":  "Hangar Storage (Y/N)",
    "ACTIVATION_DATE":     "Activation Date",
    "EFF_DATE":            "Effective Date",
}

# ── Helpers ────────────────────────────────────────────────────────────────────

def fetch_page(offset: int) -> dict:
    """Fetch a single page of records from the ArcGIS REST API."""
    params = {
        "where":         "1=1",
        "outFields":     "*",
        "returnGeometry":"false",
        "f":             "json",
        "resultOffset":  offset,
        "resultRecordCount": PAGE_SIZE,
    }
    resp = requests.get(BASE_URL, params=params, timeout=30)
    resp.raise_for_status()
    return resp.json()


def normalise_field(raw: dict, field: str):
    """Case-insensitive field lookup so minor API variations don't break us."""
    # Try exact match first
    if field in raw:
        return raw[field]
    # Try case-insensitive scan
    lower = field.lower()
    for k, v in raw.items():
        if k.lower() == lower:
            return v
    return None


def is_target_facility(attrs: dict) -> bool:
    """
    Returns True if the record is:
      • Private-use  (FACILITY_USE_CODE == 'PR')
      • AND facility name (ARPT_NAME) contains 'airpark' or 'hangar'
        (case-insensitive)
    """
    use = (normalise_field(attrs, "FACILITY_USE_CODE") or "").strip().upper()
    if use != PRIVATE_CODE:
        return False

    name = (normalise_field(attrs, "ARPT_NAME") or "").lower()
    return any(kw in name for kw in KEYWORDS)


def build_row(attrs: dict) -> list:
    """Extract the fields we want into an ordered list."""
    return [normalise_field(attrs, f) for f in FIELD_MAP]


# ── Main download loop ─────────────────────────────────────────────────────────

def download_all_facilities() -> list[list]:
    """
    Pages through the entire dataset and returns a list of matching rows.
    Each row is a list of values ordered to match FIELD_MAP.
    """
    all_rows   = []
    offset     = 0
    total_seen = 0
    page_num   = 0

    print("=" * 60)
    print("  HangarLinks — FAA Aviation Facilities Downloader")
    print("=" * 60)
    print(f"  Endpoint  : {BASE_URL}")
    print(f"  Page size : {PAGE_SIZE:,}")
    print(f"  Filter    : USE='{PRIVATE_CODE}' + keywords {KEYWORDS}")
    print("=" * 60)

    while True:
        page_num += 1
        print(f"\n[Page {page_num:>3}]  offset={offset:>6,}  "
              f"matches so far: {len(all_rows):,}", end="  …", flush=True)

        try:
            data = fetch_page(offset)
        except requests.RequestException as exc:
            print(f"\n  ⚠  HTTP error on page {page_num}: {exc}")
            print("  Retrying in 5 seconds…")
            time.sleep(5)
            continue

        features = data.get("features", [])
        if not features:
            print("  ✓ No more records — download complete.")
            break

        batch_matches = 0
        for feature in features:
            attrs = feature.get("attributes", {})
            total_seen += 1
            if is_target_facility(attrs):
                all_rows.append(build_row(attrs))
                batch_matches += 1

        print(f"fetched {len(features):,} records  |  "
              f"+{batch_matches} matches this page")

        # If the API returned fewer records than requested, we're done
        if len(features) < PAGE_SIZE:
            print("\n  ✓ Final page reached — download complete.")
            break

        offset += PAGE_SIZE
        time.sleep(REQUEST_DELAY)

    print(f"\n{'─'*60}")
    print(f"  Total records scanned : {total_seen:,}")
    print(f"  Matching facilities   : {len(all_rows):,}")
    print(f"{'─'*60}")
    return all_rows


# ── Excel export ───────────────────────────────────────────────────────────────

def write_excel(rows: list[list], filepath: str) -> None:
    """Writes the filtered rows to a styled Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "HangarLinks Facilities"

    headers = list(FIELD_MAP.values())

    # ── Header styling ──────────────────────────────────────────────────────
    header_fill   = PatternFill("solid", fgColor="1A3A5C")   # Deep navy
    accent_fill   = PatternFill("solid", fgColor="E8F0F7")   # Light blue rows
    header_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    body_font     = Font(name="Calibri", size=10)
    center        = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin_border   = Border(
        left   = Side(style="thin", color="C0C0C0"),
        right  = Side(style="thin", color="C0C0C0"),
        top    = Side(style="thin", color="C0C0C0"),
        bottom = Side(style="thin", color="C0C0C0"),
    )

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = thin_border

    ws.row_dimensions[1].height = 28

    # Write data rows
    for row_idx, row_data in enumerate(rows, start=2):
        fill = accent_fill if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = body_font
            cell.alignment = left if col_idx in (2, 8) else center
            cell.border    = thin_border
            if fill:
                cell.fill = fill

    # ── Auto-size columns (capped at 50 chars) ──────────────────────────────
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len    = len(header)
        for row_idx in range(2, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    # ── Freeze top row & add auto-filter ────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    summary_data = [
        ("HangarLinks Master List — FAA Aviation Facilities", None),
        (None, None),
        ("Generated",        "auto"),
        ("Data Source",      "FAA NTAD Aviation Facilities (ArcGIS Open Data)"),
        ("Filter — Use Type", f"Private ('{PRIVATE_CODE}')"),
        ("Filter — Keywords", ", ".join(k.title() for k in KEYWORDS)),
        ("Total Matches",    len(rows)),
    ]

    import datetime
    for r_idx, (label, value) in enumerate(summary_data, start=1):
        if label == "Generated" and value == "auto":
            value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_summary.cell(row=r_idx, column=1, value=label)
        ws_summary.cell(row=r_idx, column=2, value=value)

    ws_summary.cell(row=1, column=1).font = Font(
        name="Calibri", bold=True, size=13, color="1A3A5C"
    )
    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 55

    wb.save(filepath)
    print(f"\n  ✅  Excel file saved → {filepath}")


# ── Entry point ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    try:
        matching_rows = download_all_facilities()

        if not matching_rows:
            print("\n  ⚠  No matching facilities found. "
                  "Check your filter settings and try again.")
            sys.exit(0)

        write_excel(matching_rows, OUTPUT_FILE)

        print("\n  📋  Preview of first 5 matches:")
        print("  " + " | ".join(list(FIELD_MAP.values())[:4]))
        print("  " + "-" * 60)
        for row in matching_rows[:5]:
            print("  " + " | ".join(str(v or "—") for v in row[:4]))

        print(f"\n  Done! {len(matching_rows):,} facilities exported to "
              f"'{OUTPUT_FILE}'")

    except KeyboardInterrupt:
        print("\n\n  ⚠  Interrupted by user.")
        sys.exit(1)
